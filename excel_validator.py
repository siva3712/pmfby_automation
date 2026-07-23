from dataclasses import dataclass, field
from datetime import datetime
from typing import List

from openpyxl import load_workbook


@dataclass
class ValidationError:
    row: int
    column: str
    message: str


@dataclass
class ValidationResult:
    success: bool = True
    errors: List[ValidationError] = field(default_factory=list)

    def add(self, row, column, message):
        self.success = False
        self.errors.append(
            ValidationError(
                row=row,
                column=column,
                message=message
            )
        )


class ExcelValidator:

    REQUIRED_COLUMNS = [
        "Account No",
        "District",
        "Mandal",
        "Gram Panchayat",
        "Village",
        "Crop",
        "Khata No",
        "Survey Nos",
        "Khasra No",
        "Sowing Date",
        "Premium Debit Date"
    ]

    DATE_COLUMNS = [
        "Sowing Date",
        "Premium Debit Date"
    ]

    #######################################################

    def validate(self, file_path):

        result = ValidationResult()

        workbook = load_workbook(
            file_path,
            data_only=True
        )

        sheet = workbook.active

        ###################################################
        # Header Validation
        ###################################################

        headers = [
            str(cell.value).strip() if cell.value else ""
            for cell in sheet[1]
        ]

        for column in self.REQUIRED_COLUMNS:

            if column not in headers:

                result.add(
                    1,
                    column,
                    "Missing Required Column"
                )

        if not result.success:
            return result

        ###################################################
        # Column Mapping
        ###################################################

        index = {}

        for i, header in enumerate(headers):

            index[header] = i

        ###################################################
        # Row Validation
        ###################################################

        for row_no, row in enumerate(
                sheet.iter_rows(min_row=2),
                start=2):

            values = {}

            for column in self.REQUIRED_COLUMNS:

                value = row[
                    index[column]
                ].value

                if isinstance(value, str):
                    value = value.strip()

                values[column] = value

            ################################################
            # Mandatory Validation
            ################################################

            for column, value in values.items():

                if value is None or value == "":

                    result.add(
                        row_no,
                        column,
                        "Value is mandatory"
                    )

            ################################################
            # Survey Validation
            ################################################

            surveys = values["Survey Nos"]

            if surveys:

                survey_list = [
                    s.strip()
                    for s in str(surveys).split(",")
                ]

                for survey in survey_list:

                    if survey == "":

                        result.add(
                            row_no,
                            "Survey Nos",
                            "Empty survey number found"
                        )

            ################################################
            # Date Validation
            ################################################

            for column in self.DATE_COLUMNS:

                value = values[column]

                if value:

                    if isinstance(value, datetime):
                        continue

                    parsed = False

                    for fmt in (
                            "%d-%m-%Y",
                            "%d/%m/%Y",
                            "%Y-%m-%d"):

                        try:

                            datetime.strptime(
                                str(value),
                                fmt
                            )

                            parsed = True

                            break

                        except Exception:

                            pass

                    if not parsed:

                        result.add(
                            row_no,
                            column,
                            "Invalid Date"
                        )

            ################################################
            # Numeric Account Validation
            ################################################

            account = values["Account No"]

            if account:

                if not str(account).isdigit():

                    result.add(
                        row_no,
                        "Account No",
                        "Invalid Account Number"
                    )

        return result