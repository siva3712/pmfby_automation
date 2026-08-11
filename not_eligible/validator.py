from openpyxl import load_workbook


class NotEligibleExcelValidator:

    ALLOWED_REASONS = {

        "NPA",

        "Account-closed",

        "Crop not notified",

        "KCC-loan taken for non crop activity(s)",

        "Kcc not issued for current-season"

    }

    REQUIRED_HEADERS = {

        "Account No",

        "Reason",

        "Remarks"

    }

    def validate(self, file_path):

        errors = []

        try:

            workbook = load_workbook(
                file_path,
                read_only=True,
                data_only=True
            )

            sheet = workbook.active

            ##################################################
            # Headers
            ##################################################

            headers = []

            for cell in sheet[1]:

                value = str(
                    cell.value or ""
                ).strip()

                headers.append(value)

            header_set = set(headers)

            missing = (
                self.REQUIRED_HEADERS
                - header_set
            )

            for header in missing:

                errors.append(
                    f"Missing required column: {header}"
                )

            if missing:

                return False, errors

            ##################################################
            # Column indexes
            ##################################################

            account_index = headers.index(
                "Account No"
            )

            reason_index = headers.index(
                "Reason"
            )

            remarks_index = headers.index(
                "Remarks"
            )

            ##################################################
            # Rows
            ##################################################

            for row_number, row in enumerate(
                sheet.iter_rows(
                    min_row=2,
                    values_only=True
                ),
                start=2
            ):

                account_no = (
                    str(
                        row[account_index] or ""
                    ).strip()
                )

                reason = (
                    str(
                        row[reason_index] or ""
                    ).strip()
                )

                remarks = (
                    str(
                        row[remarks_index] or ""
                    ).strip()
                )

                ##################################################
                # Completely empty row
                ##################################################

                if not account_no and not reason and not remarks:

                    continue

                ##################################################
                # Account
                ##################################################

                if not account_no:

                    errors.append(
                        f"Row {row_number}: Account No is required."
                    )

                ##################################################
                # Reason
                ##################################################

                if not reason:

                    errors.append(
                        f"Row {row_number}: Reason is required."
                    )

                elif reason not in self.ALLOWED_REASONS:

                    errors.append(
                        f"Row {row_number}: Invalid reason '{reason}'."
                    )

                ##################################################
                # Remarks
                ##################################################

                if not remarks:

                    errors.append(
                        f"Row {row_number}: Remarks are required."
                    )

            workbook.close()

            return (
                len(errors) == 0,
                errors
            )

        except Exception as ex:

            return False, [
                f"Unable to validate Excel: {str(ex)}"
            ]