from openpyxl import load_workbook

from .models import NotEligibleAccount


class NotEligibleExcelReader:

    def read(self, file_path):

        workbook = load_workbook(
            file_path,
            read_only=True,
            data_only=True
        )

        sheet = workbook.active

        ##################################################
        # Read headers
        ##################################################

        headers = [
            str(cell.value or "").strip()
            for cell in sheet[1]
        ]

        account_index = headers.index(
            "Account No"
        )

        reason_index = headers.index(
            "Reason"
        )

        remarks_index = headers.index(
            "Remarks"
        )

        accounts = []

        ##################################################
        # Read rows
        ##################################################

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            account_no = str(
                row[account_index] or ""
            ).strip()

            reason = str(
                row[reason_index] or ""
            ).strip()

            remarks = str(
                row[remarks_index] or ""
            ).strip()

            ##################################################
            # Ignore completely empty rows
            ##################################################

            if not account_no:
                continue

            accounts.append(
                NotEligibleAccount(
                    account_no=account_no,
                    reason=reason,
                    remarks=remarks
                )
            )

        workbook.close()

        return accounts