from datetime import datetime, date

from openpyxl import load_workbook

from models import Account, Khata, Crop


class ExcelReader:

    ########################################################

    def read(self, file_path):

        workbook = load_workbook(
            file_path,
            data_only=True
        )

        sheet = workbook.active

        headers = [
            str(cell.value).strip()
            for cell in sheet[1]
        ]

        column = {}

        for i, header in enumerate(headers):
            column[header] = i

        accounts = {}

        ####################################################
        # Read every row
        ####################################################

        for row in sheet.iter_rows(min_row=2):

            account_no = str(
                row[column["Account No"]].value
            ).strip()

            district = str(
                row[column["District"]].value
            ).strip()

            mandal = str(
                row[column["Mandal"]].value
            ).strip()

            gram_panchayat = str(
                row[column["Gram Panchayat"]].value
            ).strip()

            village = str(
                row[column["Village"]].value
            ).strip()

            khata_no = str(
                row[column["Khata No"]].value
            ).strip()

            khasra_no = str(
                row[column["Khasra No"]].value
            ).strip()

            crop_name = str(
                row[column["Crop"]].value
            ).strip()

            survey_numbers = self.parse_surveys(
                row[column["Survey Nos"]].value
            )

            sowing_date = self.parse_date(
                row[column["Sowing Date"]].value
            )

            premium_debit_date = self.parse_date(
                row[column["Premium Debit Date"]].value
            )

            ################################################
            # Account
            ################################################

            if account_no not in accounts:

                accounts[account_no] = Account(
                    account_no=account_no,
                    district=district,
                    mandal=mandal,
                    gram_panchayat=gram_panchayat,
                    village=village
                )

            account = accounts[account_no]

            ################################################
            # Khata
            ################################################

            khata = account.find_khata(
                khata_no
            )

            if khata is None:

                khata = Khata(
                    khata_no=khata_no,
                    khasra_no=khasra_no
                )

                account.add_khata(
                    khata
                )

            ################################################
            # Crop
            ################################################

            crop = khata.find_crop(
                crop_name
            )

            if crop is None:

                crop = Crop(
                    crop_name=crop_name,
                    sowing_date=sowing_date,
                    premium_debit_date=premium_debit_date,
                    survey_numbers=survey_numbers
                )

                khata.add_crop(
                    crop
                )

            else:

                for survey in survey_numbers:

                    if survey not in crop.survey_numbers:

                        crop.survey_numbers.append(
                            survey
                        )

        return list(accounts.values())

    ########################################################

    def parse_surveys(self, surveys):

        if surveys is None:

            return []

        return [
            s.strip()
            for s in str(surveys).split(",")
            if s.strip()
        ]

    ########################################################

    def parse_date(self, value):

        if isinstance(value, datetime):

            return value.date()

        if isinstance(value, date):

            return value

        for fmt in (
                "%d-%m-%Y",
                "%d/%m/%Y",
                "%Y-%m-%d"):

            try:

                return datetime.strptime(
                    str(value),
                    fmt
                ).date()

            except Exception:

                pass

        return None