from openpyxl import load_workbook

from .change_account import ChangeAccount


class ChangeAccountExcelReader:

    def read(self, file_path):

        workbook = load_workbook(
            file_path,
            data_only=True
        )

        sheet = workbook.active

        accounts = []

        headers = [
            cell.value
            for cell in sheet[1]
        ]

        account_index = headers.index(
            "Account Number"
        )

        new_account_index = headers.index(
            "New Account Number"
        )

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            if not row:
                continue

            account_no = row[
                account_index
            ]

            new_account_no = row[
                new_account_index
            ]

            if (
                account_no is None
                and new_account_no is None
            ):
                continue

            accounts.append(

                ChangeAccount(

                    account_no=account_no,

                    new_account_no=new_account_no

                )

            )

        return accounts