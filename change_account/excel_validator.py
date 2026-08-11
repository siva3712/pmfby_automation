from openpyxl import load_workbook


class ChangeAccountExcelValidator:

    REQUIRED_COLUMNS = [

        "Account Number",

        "New Account Number"

    ]

    MAX_ACCOUNT_LENGTH = 16

    ##########################################################

    def validate(self, excel_file):

        errors = []

        workbook = None

        try:

            ##################################################
            # Open workbook
            ##################################################

            workbook = load_workbook(

                excel_file,

                read_only=True,

                data_only=True

            )

            sheet = workbook.active

            ##################################################
            # Check header
            ##################################################

            header_row = next(
                sheet.iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True
                )
            )

            headers = []

            for value in header_row:

                if value is None:

                    headers.append("")

                else:

                    headers.append(
                        str(value).strip()
                    )

            ##################################################
            # Required columns
            ##################################################

            column_indexes = {}

            for column in self.REQUIRED_COLUMNS:

                if column not in headers:

                    errors.append(

                        f"Missing required column: {column}"

                    )

                else:

                    column_indexes[column] = (
                        headers.index(column)
                    )

            ##################################################
            # Cannot continue without headers
            ##################################################

            if errors:

                return False, errors

            account_index = column_indexes[
                "Account Number"
            ]

            new_account_index = column_indexes[
                "New Account Number"
            ]

            ##################################################
            # Validate rows
            ##################################################

            for row_number, row in enumerate(

                sheet.iter_rows(
                    min_row=2,
                    values_only=True
                ),

                start=2

            ):

                ##################################################
                # Ignore completely empty rows
                ##################################################

                if all(
                    value is None
                    for value in row
                ):

                    continue

                ##################################################
                # Old account
                ##################################################

                account_no = self._clean_account(
                    row[account_index]
                    if account_index < len(row)
                    else None
                )

                ##################################################
                # New account
                ##################################################

                new_account_no = self._clean_account(
                    row[new_account_index]
                    if new_account_index < len(row)
                    else None
                )

                ##################################################
                # Old account validation
                ##################################################

                if not account_no:

                    errors.append(

                        f"Row {row_number} | "
                        f"Account Number | "
                        f"Account Number is required."

                    )

                else:

                    if not account_no.isdigit():

                        errors.append(

                            f"Row {row_number} | "
                            f"Account Number | "
                            f"Account Number must contain only digits."

                        )

                    elif len(account_no) > self.MAX_ACCOUNT_LENGTH:

                        errors.append(

                            f"Row {row_number} | "
                            f"Account Number | "
                            f"Maximum {self.MAX_ACCOUNT_LENGTH} digits allowed."

                        )

                ##################################################
                # New account validation
                ##################################################

                if not new_account_no:

                    errors.append(

                        f"Row {row_number} | "
                        f"New Account Number | "
                        f"New Account Number is required."

                    )

                else:

                    if not new_account_no.isdigit():

                        errors.append(

                            f"Row {row_number} | "
                            f"New Account Number | "
                            f"New Account Number must contain only digits."

                        )

                    elif len(new_account_no) > self.MAX_ACCOUNT_LENGTH:

                        errors.append(

                            f"Row {row_number} | "
                            f"New Account Number | "
                            f"Maximum {self.MAX_ACCOUNT_LENGTH} digits allowed."

                        )

                ##################################################
                # Old and new cannot be same
                ##################################################

                if (
                    account_no
                    and new_account_no
                    and account_no == new_account_no
                ):

                    errors.append(

                        f"Row {row_number} | "
                        f"New Account Number | "
                        f"New Account Number must be different from "
                        f"the existing Account Number."

                    )

            ##################################################
            # Final result
            ##################################################

            return (

                len(errors) == 0,

                errors

            )

        finally:

            if workbook:

                workbook.close()

    ##########################################################

    @staticmethod
    def _clean_account(value):

        if value is None:

            return ""

        ##################################################
        # Excel may return numeric cells as integers
        ##################################################

        if isinstance(value, int):

            return str(value)

        ##################################################
        # Excel may return float for numeric cells
        # e.g. 19089620932.0
        ##################################################

        if isinstance(value, float):

            if value.is_integer():

                return str(
                    int(value)
                )

        ##################################################
        # Normal string
        ##################################################

        return str(value).strip()